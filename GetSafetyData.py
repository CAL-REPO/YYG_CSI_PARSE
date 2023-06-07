# Python Version 3.11

#Library List
"""
python.exe -m pip install --upgrade pip
pip3 install selenium
pip3 install webdriver-manager
pip3 install pywin32
pip3 install requests
pip3 install bs4
pip3 install lxml
pip3 install pandas
pip3 install openpyxl
"""

import os
import re
import requests
import pandas as pd
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from bs4 import BeautifulSoup, SoupStrainer
from time import sleep
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager

# Set Selenium Option
options = webdriver.ChromeOptions()
options.add_argument("--ignore-certificate-error")
options.add_argument("--ignore-ssl-errors")
options.add_argument("--disable-web-security")
options.add_argument("--allow-running-insecure-content")
options.add_experimental_option("excludeSwitches", ["enable-logging"])
options.add_experimental_option("excludeSwitches", ["enable-automation"])

acdMainUrl = "https://www.csi.go.kr/acd/acdCaseList.do"
acdCaseUrl = "https://www.csi.go.kr/acd/acdCaseView.do?case_no="

# Check if User Input CasePageNo or not
InputEndCasePageNo = input("Enter a End Case Page No: ")
if InputEndCasePageNo != "" :
    SearchPageCount = 2
    InputEndCasePageNo = int(InputEndCasePageNo)
else:
    SearchPageCount = 11
    InputEndCasePageNo = str(InputEndCasePageNo)

# Get Recent Accident Example Page Number
## Execute Selenium
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
driver.maximize_window()
rctCasePageNoList = []

for xPathtrNo in range(1, SearchPageCount):

    driver.get(acdMainUrl)

    ## Get Recent Accident Example HTML Element from Main Page
    IDforWait = "main"
    WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.ID,IDforWait)))

    rctCaseXPath = "/html/body/div[3]/div[3]/div/div/div[3]/section/form/div/div[1]/table/tbody/tr[" + str(xPathtrNo) + "]/td[1]/a"
    rctCaseText = driver.find_element(By.XPATH,rctCaseXPath).text
    rctCaseEle = "//*[text()='" + rctCaseText + "']"
    rctCaseJsEle = driver.find_element(By.XPATH,rctCaseEle)

    ## Move to Recent Accident Example Page
    driver.execute_script("arguments[0].click();", rctCaseJsEle)

    IDforWait = "content"
    WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.ID,IDforWait)))

    rctCasePageNoXPath = "/html/body/div[3]/div[3]/div/div/div[3]/section/div[1]"
    rctCasePageNoText = driver.find_element(By.XPATH, rctCasePageNoXPath).text
    rctCasePageNoTextList = re.findall(r'\d+', rctCasePageNoText)
    rctCasePageNoList.append(rctCasePageNoTextList[0])

if type(InputEndCasePageNo) is int:
    rctCasePageNo = InputEndCasePageNo
else:
    rctCasePageNo = int(max(rctCasePageNoList))
rctCaseUrl = acdCaseUrl + str(rctCasePageNo)
## Quit Selenium Session
driver.quit()

# Parse Safety Data Header
response = requests.get(rctCaseUrl)
content = SoupStrainer(id = "main")
soup = BeautifulSoup(response.content.decode('utf-8'), 'html.parser', parse_only=content) 
header  = [str(td.get_text().strip()) for td in soup.find_all('td', class_ = "td-head", attrs = "") if str(td) and not td.has_attr('rowspan')] 
header.insert(0 , "PageNo")

# Create Safety Data Dictionary
dictData = {}
for dictKeyIndex in range(0, len(header)):
    dictData[header[dictKeyIndex]] = []

# ##### Set csv File
##### cwd = os.getcwd()
##### csvFileName = "SafetyData.csv"
##### csvFileFullPath = cwd + "\\" + csvFileName

# Set xlsx File
cwd = os.getcwd()
xlFileName = "SafetyData.xlsx"
xlFileFullPath = cwd + "\\" + xlFileName
wb = openpyxl.load_workbook(xlFileFullPath)
ws = wb.worksheets[0]
if not ws.cell(1,1).value == 'PageNo' : ws.cell(1,1).value = 'PageNo'
startRow = ws.max_row
wb.save(xlFileFullPath)
wb.close()

# ##### Check Recent Accident Case PageNo in Safety Data csv File
##### dfCsvData = pd.read_csv(csvFileFullPath, na_values='', sep=',', header=0, usecols=['PageNo'], encoding='cp949')

# Check Recent Accident Case PageNo in Safety Data xlsx File
dfXlData = pd.read_excel(xlFileFullPath, keep_default_na=False, header=0, usecols=['PageNo'])

# Change Data Array Type from DataFrame to List
##### listCsvData = dfCsvData.values.tolist()
##### fileRctCasePageNo = "" if len(listCsvData) == 0 else listCsvData[0][0]

# Change Data Array Type from DataFrame to List
listXlData = dfXlData.values.tolist()
fileRctCasePageNo = "" if len(listXlData) == 0 else listXlData[0][0]
del dfXlData
del listXlData

# Parse Safety Data Row
iniCasePageNo = 0 if fileRctCasePageNo == "" else fileRctCasePageNo
CutSize = 20

# Check if Recent Case Page No is smller than Initial Page No or not
if rctCasePageNo > iniCasePageNo :
    for casePageCutStart in range(rctCasePageNo, iniCasePageNo, -CutSize):
        casePageCutEnd = max(casePageCutStart - CutSize, iniCasePageNo)

        for casePageNo in range(casePageCutStart, casePageCutEnd, -1):
            caseUrl = acdCaseUrl + str(casePageNo)
            response = requests.get(caseUrl)
            content = SoupStrainer(id = "main")
            soup = BeautifulSoup(response.content.decode('utf-8'), 'html.parser', parse_only=content)
            data = [ILLEGAL_CHARACTERS_RE.sub(r'', str(td.get_text().replace('\r',' ').replace('\n','').replace('\xa0','').replace('\u30fb','').replace('\u2013','').replace('\u2013','').strip())) for td in soup.find_all('td', class_ = "t-left", attrs = "") if str(td)]
            data.insert(0, str(casePageNo))
            print(casePageNo)
            for dictIndex in range(0, len(dictData)):
                dictData[str(header[dictIndex])].append(str(data[dictIndex]))
            
        # Change Data Format from Dictionary to Pandas DataFrame
        dfData = pd.DataFrame.from_dict(dictData)

        # Input Safety Data to csv file
        if fileRctCasePageNo == "":
            ##### dfData.to_csv(csvFileFullPath, na_rep='NULL', sep=',', index=False, header=True, encoding='cp949')
            dfData.to_excel(xlFileFullPath, na_rep='NULL', index=False, header=True)

        else:
            ##### dfData.to_csv(csvFileFullPath, mode='a', na_rep='NULL', sep=',', index=False, header=False, encoding='cp949')
            with pd.ExcelWriter(xlFileFullPath, engine='openpyxl', mode='a', if_sheet_exists='overlay') as xlWriter:
                dfData.to_excel(xlWriter, startrow=startRow ,na_rep='NULL', index=False, header=False)
    
        del dfData
        sleep(1)
    # Sort data in Desending Order By CasePageNo
    dfResult = pd.read_excel(xlFileFullPath, keep_default_na=False)
    dfSort = dfResult.sort_values(by = "PageNo", ascending=False)
    dfSort.to_excel(xlFileFullPath,na_rep='NULL', index=False, header=True)
    del dfResult
    del dfSort